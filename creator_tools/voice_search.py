from __future__ import annotations

import argparse
import csv
from functools import lru_cache
import json
import math
import os
import subprocess
import sys
import tempfile
from pathlib import Path

import numpy as np
import soundfile as sf


FORK_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_CREATOR_ROOT = (
    FORK_ROOT.parents[1]
    if FORK_ROOT.parent.name == "third_party"
    else FORK_ROOT
)
PROJECT_ROOT = Path(
    os.environ.get("CREATOR_ROOT", str(DEFAULT_CREATOR_ROOT))
).resolve()
TOOL_ROOT = PROJECT_ROOT / "output" / "data" / "tools" / "emotivoice"
VOICE_SEARCH_ROOT = TOOL_ROOT / "voice-search"
DEFAULT_MODEL = (
    VOICE_SEARCH_ROOT
    / "models"
    / "3dspeaker_speech_campplus_sv_zh-cn_16k-common.onnx"
)
DEFAULT_INDEX = VOICE_SEARCH_ROOT / "index" / "emotivoice-speakers.npz"
DEFAULT_METADATA = FORK_ROOT / "creator_tools" / "data" / "speaker_metadata.csv"
DEFAULT_SAMPLES = VOICE_SEARCH_ROOT / "reference-samples"
DEFAULT_VIDEOS = PROJECT_ROOT / "output" / "data" / "douyin" / "videos"
DEFAULT_REPORT = (
    PROJECT_ROOT / "output" / "data" / "tools" / "emotivoice" / "作者音色匹配.xlsx"
)
SAMPLE_RATE = 16000
SEGMENT_SECONDS = 12.0
SEGMENT_POSITIONS = (0.18, 0.5, 0.82)


def log(message: str) -> None:
    print(message, file=sys.stderr, flush=True)


def normalize(vector: np.ndarray) -> np.ndarray:
    vector = np.asarray(vector, dtype=np.float32).reshape(-1)
    norm = float(np.linalg.norm(vector))
    if not math.isfinite(norm) or norm <= 0:
        raise ValueError("Speaker embedding has an invalid norm.")
    return vector / norm


def load_metadata(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as file:
        return list(csv.DictReader(file))


def write_metadata(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def create_extractor(model_path: Path, num_threads: int):
    import sherpa_onnx

    config = sherpa_onnx.SpeakerEmbeddingExtractorConfig(
        model=str(model_path),
        num_threads=num_threads,
        provider="cpu",
    )
    if not config.validate():
        raise ValueError(f"Invalid speaker embedding config: {config}")
    return sherpa_onnx.SpeakerEmbeddingExtractor(config)


def embedding_from_samples(extractor, samples: np.ndarray, sample_rate: int):
    samples = np.ascontiguousarray(samples, dtype=np.float32)
    if samples.size < sample_rate // 2:
        raise ValueError("Audio must contain at least 0.5 seconds of speech.")
    stream = extractor.create_stream()
    stream.accept_waveform(sample_rate=sample_rate, waveform=samples)
    stream.input_finished()
    if not extractor.is_ready(stream):
        raise RuntimeError("Speaker embedding extractor is not ready.")
    return normalize(np.asarray(extractor.compute(stream), dtype=np.float32))


def embedding_from_wav(extractor, path: Path) -> np.ndarray:
    samples, sample_rate = sf.read(path, always_2d=True, dtype="float32")
    return embedding_from_samples(extractor, samples[:, 0], sample_rate)


def decode_media(path: Path, output: Path) -> None:
    ffmpeg = os.environ.get("FFMPEG_BINARY", "ffmpeg")
    result = subprocess.run(
        [
            ffmpeg,
            "-hide_banner",
            "-loglevel",
            "error",
            "-y",
            "-i",
            str(path),
            "-vn",
            "-ac",
            "1",
            "-ar",
            str(SAMPLE_RATE),
            "-c:a",
            "pcm_s16le",
            "-af",
            "highpass=f=70,lowpass=f=7800",
            str(output),
        ],
        check=False,
        capture_output=True,
        text=True,
    )
    if result.returncode:
        details = result.stderr.strip() or f"退出代码 {result.returncode}"
        raise ValueError(f"FFmpeg 无法解码该音视频：{details}")


def embedding_from_media(extractor, path: Path) -> np.ndarray:
    with tempfile.TemporaryDirectory(prefix="emotivoice-search-") as temp_dir:
        wav_path = Path(temp_dir) / "decoded.wav"
        decode_media(path, wav_path)
        samples, sample_rate = sf.read(
            wav_path,
            always_2d=True,
            dtype="float32",
        )

    mono = samples[:, 0]
    duration = len(mono) / sample_rate
    if duration <= 0:
        raise ValueError(f"Audio duration is invalid: {path}")
    if duration <= SEGMENT_SECONDS * 1.5:
        segments = [(0, len(mono))]
    else:
        segment_size = int(SEGMENT_SECONDS * sample_rate)
        segments = []
        for position in SEGMENT_POSITIONS:
            center = int(len(mono) * position)
            start = max(0, min(len(mono) - segment_size, center - segment_size // 2))
            segments.append((start, start + segment_size))

    embeddings = [
        embedding_from_samples(extractor, mono[start:end], sample_rate)
        for start, end in segments
    ]
    return normalize(np.mean(np.stack(embeddings), axis=0))


def estimate_pitch(samples: np.ndarray, sample_rate: int) -> float:
    frame_size = int(sample_rate * 0.04)
    hop_size = int(sample_rate * 0.02)
    min_lag = max(1, int(sample_rate / 400))
    max_lag = min(frame_size - 1, int(sample_rate / 65))
    pitches = []
    for start in range(0, max(0, len(samples) - frame_size), hop_size):
        frame = samples[start : start + frame_size].astype(np.float64)
        frame -= frame.mean()
        rms = float(np.sqrt(np.mean(frame * frame)))
        if rms < 0.01:
            continue
        frame *= np.hanning(frame_size)
        spectrum = np.fft.rfft(frame, n=frame_size * 2)
        autocorrelation = np.fft.irfft(
            spectrum * np.conjugate(spectrum)
        )[:frame_size]
        if autocorrelation[0] <= 0:
            continue
        lag = min_lag + int(
            np.argmax(autocorrelation[min_lag : max_lag + 1])
        )
        confidence = autocorrelation[lag] / autocorrelation[0]
        if confidence >= 0.25:
            pitches.append(sample_rate / lag)
    return float(np.median(pitches)) if pitches else 0.0


def acoustic_metrics(path: Path) -> tuple[float, float]:
    samples, sample_rate = sf.read(path, always_2d=True, dtype="float32")
    mono = samples[:, 0]
    duration = len(mono) / sample_rate
    return estimate_pitch(mono, sample_rate), duration


def percentile_thresholds(
    rows: list[dict[str, str]],
    metrics: dict[str, tuple[float, float]],
    gender: str,
) -> tuple[float, float]:
    pitches = [
        metrics[row["speaker_id"]][0]
        for row in rows
        if row["gender"] == gender and metrics[row["speaker_id"]][0] > 0
    ]
    low, high = np.quantile(np.asarray(pitches), [0.33, 0.67])
    return float(low), float(high)


def enrich_metadata(
    rows: list[dict[str, str]],
    metrics: dict[str, tuple[float, float]],
) -> None:
    pitch_thresholds = {
        gender: percentile_thresholds(rows, metrics, gender)
        for gender in ("F", "M")
    }
    durations = np.asarray([duration for _, duration in metrics.values()])
    duration_low, duration_high = np.quantile(durations, [0.33, 0.67])

    for row in rows:
        pitch, duration = metrics[row["speaker_id"]]
        low, high = pitch_thresholds[row["gender"]]
        if row["gender"] == "F":
            tone = "沉稳" if pitch < low else "柔和" if pitch < high else "清亮"
        else:
            tone = "低沉" if pitch < low else "温和" if pitch < high else "明朗"
        pace = (
            "利落"
            if duration < duration_low
            else "自然"
            if duration < duration_high
            else "舒缓"
        )
        serial = row["chinese_name"][-4:]
        row["pitch_hz"] = f"{pitch:.1f}"
        row["duration_s"] = f"{duration:.2f}"
        row["voice_profile"] = f"{tone}、{pace}"
        row["chinese_name"] = f"{row['gender_zh']}·{tone}{pace}·{serial}"
        row["display_name"] = (
            f"{row['chinese_name']}｜{row['english_name']}｜ID {row['speaker_id']}"
        )


def build_index(args) -> None:
    metadata = load_metadata(args.metadata)
    extractor = create_extractor(args.model, args.num_threads)
    speaker_ids = []
    embeddings = []
    metrics: dict[str, tuple[float, float]] = {}

    for index, row in enumerate(metadata, start=1):
        speaker_id = row["speaker_id"]
        sample_path = args.samples / f"{speaker_id}.wav"
        if not sample_path.is_file():
            raise FileNotFoundError(f"Reference sample is missing: {sample_path}")
        embeddings.append(embedding_from_wav(extractor, sample_path))
        metrics[speaker_id] = acoustic_metrics(sample_path)
        speaker_ids.append(speaker_id)
        if index % 50 == 0 or index == len(metadata):
            log(f"Indexed {index}/{len(metadata)} reference voices.")

    enrich_metadata(metadata, metrics)
    write_metadata(args.metadata, metadata)
    args.index.parent.mkdir(parents=True, exist_ok=True)
    np.savez_compressed(
        args.index,
        speaker_ids=np.asarray(speaker_ids),
        embeddings=np.stack(embeddings).astype(np.float32),
    )
    print(f"Built {len(speaker_ids)}-voice index at {args.index}")


def load_index(path: Path) -> tuple[np.ndarray, np.ndarray]:
    with np.load(path) as data:
        return data["speaker_ids"], data["embeddings"]


def search_embedding(
    query: np.ndarray,
    speaker_ids: np.ndarray,
    embeddings: np.ndarray,
    metadata: list[dict[str, str]],
    top_k: int,
) -> list[dict[str, object]]:
    metadata_by_id = {row["speaker_id"]: row for row in metadata}
    scores = embeddings @ normalize(query)
    top_indices = np.argsort(scores)[::-1][:top_k]
    results = []
    for rank, index in enumerate(top_indices, start=1):
        speaker_id = str(speaker_ids[index])
        row = metadata_by_id[speaker_id]
        results.append(
            {
                "rank": rank,
                "speaker_id": speaker_id,
                "chinese_name": row["chinese_name"],
                "english_name": row["english_name"],
                "gender": row["gender_zh"],
                "voice_profile": row["voice_profile"],
                "similarity": round(float(scores[index]), 4),
                "sample_path": str(DEFAULT_SAMPLES / f"{speaker_id}.wav"),
            }
        )
    return results


@lru_cache(maxsize=2)
def load_search_resources(
    model_path: str,
    index_path: str,
    metadata_path: str,
    num_threads: int,
):
    metadata = load_metadata(Path(metadata_path))
    speaker_ids, embeddings = load_index(Path(index_path))
    extractor = create_extractor(Path(model_path), num_threads)
    return metadata, speaker_ids, embeddings, extractor


def find_similar_voices(
    input_path: Path,
    *,
    model_path: Path = DEFAULT_MODEL,
    index_path: Path = DEFAULT_INDEX,
    metadata_path: Path = DEFAULT_METADATA,
    top_k: int = 8,
    num_threads: int = 4,
) -> list[dict[str, object]]:
    required = [model_path, index_path, metadata_path]
    missing = [str(path) for path in required if not path.is_file()]
    if missing:
        raise FileNotFoundError(
            "声音检索文件不完整：\n" + "\n".join(missing)
        )
    metadata, speaker_ids, embeddings, extractor = load_search_resources(
        str(model_path.resolve()),
        str(index_path.resolve()),
        str(metadata_path.resolve()),
        num_threads,
    )
    query = embedding_from_media(extractor, input_path)
    return search_embedding(
        query,
        speaker_ids,
        embeddings,
        metadata,
        top_k,
    )


def search_audio(args) -> list[dict[str, object]]:
    results = find_similar_voices(
        args.input,
        model_path=args.model,
        index_path=args.index,
        metadata_path=args.metadata,
        top_k=args.top_k,
        num_threads=args.num_threads,
    )
    if args.json:
        print(json.dumps(results, ensure_ascii=False))
    else:
        print("排名\t相似度\tSpeaker ID\t性别\t中文名称\t官方英文名")
        for row in results:
            print(
                f"{row['rank']}\t{row['similarity']:.4f}\t"
                f"{row['speaker_id']}\t{row['gender']}\t"
                f"{row['chinese_name']}\t{row['english_name']}"
            )
    return results


def choose_author_videos(author_dir: Path, limit: int) -> list[Path]:
    videos = list(author_dir.rglob("*.mp4"))
    return sorted(videos, key=lambda path: path.stat().st_size, reverse=True)[:limit]


def style_workbook(path: Path, summary_rows, candidate_rows, notes) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    summary = workbook.active
    summary.title = "匹配结果"
    summary.append(
        [
            "作者",
            "参考视频数",
            "参考作品ID",
            "Speaker ID",
            "中文音色名",
            "性别",
            "官方英文名",
            "声音特征",
            "相似度",
            "第二候选",
            "第三候选",
            "状态",
        ]
    )
    for row in summary_rows:
        summary.append(row)

    candidates = workbook.create_sheet("候选详情")
    candidates.append(
        [
            "作者",
            "排名",
            "Speaker ID",
            "中文音色名",
            "性别",
            "官方英文名",
            "声音特征",
            "相似度",
        ]
    )
    for row in candidate_rows:
        candidates.append(row)

    explanation = workbook.create_sheet("说明")
    explanation.append(["项目", "内容"])
    for title, content in notes:
        explanation.append([title, content])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column in sheet.columns:
            width = min(
                60,
                max(10, max(len(str(cell.value or "")) for cell in column) + 2),
            )
            sheet.column_dimensions[get_column_letter(column[0].column)].width = width
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def match_authors(args) -> None:
    metadata = load_metadata(args.metadata)
    speaker_ids, embeddings = load_index(args.index)
    extractor = create_extractor(args.model, args.num_threads)
    summary_rows = []
    candidate_rows = []

    author_dirs = sorted(
        [path for path in args.videos_root.iterdir() if path.is_dir()],
        key=lambda path: path.name.casefold(),
    )
    if args.max_authors:
        author_dirs = author_dirs[: args.max_authors]

    for author_index, author_dir in enumerate(author_dirs, start=1):
        videos = choose_author_videos(author_dir, args.videos_per_author)
        log(f"[{author_index}/{len(author_dirs)}] Matching {author_dir.name}...")
        author_embeddings = []
        used_videos = []
        errors = []
        for video in videos:
            try:
                author_embeddings.append(embedding_from_media(extractor, video))
                used_videos.append(video)
            except (subprocess.CalledProcessError, ValueError, RuntimeError) as error:
                errors.append(f"{video.name}: {error}")

        if not author_embeddings:
            summary_rows.append(
                [
                    author_dir.name,
                    0,
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "; ".join(errors) or "没有本地 MP4",
                ]
            )
            continue

        query = normalize(np.mean(np.stack(author_embeddings), axis=0))
        results = search_embedding(
            query,
            speaker_ids,
            embeddings,
            metadata,
            args.top_k,
        )
        best = results[0]
        second = results[1] if len(results) > 1 else None
        third = results[2] if len(results) > 2 else None
        summary_rows.append(
            [
                author_dir.name,
                len(used_videos),
                "、".join(video.parent.name for video in used_videos),
                best["speaker_id"],
                best["chinese_name"],
                best["gender"],
                best["english_name"],
                best["voice_profile"],
                best["similarity"],
                (
                    f"{second['speaker_id']}｜{second['chinese_name']}｜"
                    f"{second['similarity']:.4f}"
                    if second
                    else ""
                ),
                (
                    f"{third['speaker_id']}｜{third['chinese_name']}｜"
                    f"{third['similarity']:.4f}"
                    if third
                    else ""
                ),
                "成功" if not errors else f"成功；跳过 {len(errors)} 个异常视频",
            ]
        )
        for candidate in results:
            candidate_rows.append(
                [
                    author_dir.name,
                    candidate["rank"],
                    candidate["speaker_id"],
                    candidate["chinese_name"],
                    candidate["gender"],
                    candidate["english_name"],
                    candidate["voice_profile"],
                    candidate["similarity"],
                ]
            )

    notes = [
        ("方法", "固定中文文本生成 2,014 个 EmotiVoice 参考音色，再用中文 CAM++ 声纹向量计算余弦相似度。"),
        ("作者参考", f"每位作者最多选择 {args.videos_per_author} 个本地视频，并从前、中、后位置抽取音频后平均。"),
        ("分数解释", "相似度只用于候选排序，不代表同一人或身份确认；跨语言、背景音乐、混响和真人/合成域差异都会影响结果。"),
        ("建议", "优先试听前三名；如果作者视频包含多位说话人，应换成单人、少背景音乐的 5-30 秒片段重新搜索。"),
        ("模型", "3D-Speaker CAM++ 中文通用模型，经 sherpa-onnx 离线运行。"),
    ]
    style_workbook(args.output, summary_rows, candidate_rows, notes)

    csv_path = args.output.with_suffix(".csv")
    with csv_path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(
            [
                "作者",
                "参考视频数",
                "参考作品ID",
                "Speaker ID",
                "中文音色名",
                "性别",
                "官方英文名",
                "声音特征",
                "相似度",
                "第二候选",
                "第三候选",
                "状态",
            ]
        )
        writer.writerows(summary_rows)
    print(f"Wrote {len(summary_rows)} author matches to {args.output}")


def add_common_paths(parser) -> None:
    parser.add_argument("--model", type=Path, default=DEFAULT_MODEL)
    parser.add_argument("--metadata", type=Path, default=DEFAULT_METADATA)
    parser.add_argument("--index", type=Path, default=DEFAULT_INDEX)
    parser.add_argument("--num-threads", type=int, default=4)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build and query a local EmotiVoice speaker similarity index."
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    index_parser = subparsers.add_parser("index")
    add_common_paths(index_parser)
    index_parser.add_argument("--samples", type=Path, default=DEFAULT_SAMPLES)
    index_parser.set_defaults(handler=build_index)

    search_parser = subparsers.add_parser("search")
    add_common_paths(search_parser)
    search_parser.add_argument("input", type=Path)
    search_parser.add_argument("--top-k", type=int, default=8)
    search_parser.add_argument("--json", action="store_true")
    search_parser.set_defaults(handler=search_audio)

    match_parser = subparsers.add_parser("match-authors")
    add_common_paths(match_parser)
    match_parser.add_argument("--videos-root", type=Path, default=DEFAULT_VIDEOS)
    match_parser.add_argument("--output", type=Path, default=DEFAULT_REPORT)
    match_parser.add_argument("--videos-per-author", type=int, default=3)
    match_parser.add_argument("--top-k", type=int, default=5)
    match_parser.add_argument("--max-authors", type=int)
    match_parser.set_defaults(handler=match_authors)

    args = parser.parse_args()
    args.handler(args)


if __name__ == "__main__":
    main()
